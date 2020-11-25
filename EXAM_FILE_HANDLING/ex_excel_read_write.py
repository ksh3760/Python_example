# EXCEL 파일 Read & Write
# ----------------------------------------------
from openpyxl import Workbook, load_workbook
#                   (function)   (class)

# 파일명 선언
E_FILE = './DATA/sample.xlsx'

#  EXCEL 파일 Write ---------------------------------
# 엑셀 파일 생성
wb = Workbook()

# 시트에 데이터 기록
ws = wb.active
ws.title = 'Address'
ws['A1'] = "Language"
ws['A2'] = "C"
ws['A3'] = "JAVA"
ws['A4'] = "Python"

ws['B1'] = "Creator"
ws['B2'] = "C**"
ws['B3'] = "J**"
ws['B4'] = "P**"

# 저장하기
wb.save(E_FILE)

# EXCEL 파일 Read ---------------------------------
# 1) 엑셀 파일 열기
book = load_workbook(E_FILE)

# 2) 데이터 추출
sheet = book.worksheets[0]
print(f"sheet.title : {sheet.title}")
print(f"sheet.max_row : {sheet.max_row}")
print(f"sheet.max_column : {sheet.max_column}")
print(f"sheet.rows : {sheet.rows}")

for row in sheet.rows:
    print(f"row => {row}, {type(row)}")
    print(f"row => {row[0].value}, {row[1].value}")
